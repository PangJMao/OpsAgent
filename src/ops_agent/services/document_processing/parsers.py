from __future__ import annotations

import csv
import html
import importlib
import re
import zipfile
from abc import ABC, abstractmethod
from html.parser import HTMLParser
from pathlib import Path
from xml.etree import ElementTree

from ops_agent.config import settings
from ops_agent.services.document_processing.cleaners import clean_text
from ops_agent.services.document_processing.schema import ParsedBlock, ParsedDocument


class DocumentParser(ABC):
    parser_name = "base"

    @abstractmethod
    def parse(self, path: Path) -> ParsedDocument:
        ...


class MarkdownParser(DocumentParser):
    parser_name = "markdown"

    def parse(self, path: Path) -> ParsedDocument:
        text = path.read_text(encoding="utf-8")
        return _document(path, "markdown", self.parser_name, [ParsedBlock(text=text, block_type="markdown")])


class TextParser(DocumentParser):
    parser_name = "text"

    def parse(self, path: Path) -> ParsedDocument:
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = path.read_text(encoding="gb18030", errors="replace")
        return _document(path, "txt", self.parser_name, [ParsedBlock(text=text, block_type="text")])


class PdfParser(DocumentParser):
    parser_name = "pypdf"

    def parse(self, path: Path) -> ParsedDocument:
        backend = settings.pdf_parser_backend.lower()
        if backend in {"auto", "docling", "unstructured", "mineru"}:
            parsed = _try_complex_pdf_backend(path, backend)
            if parsed is not None:
                return parsed

        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("PDF parsing requires pypdf.") from exc

        reader = PdfReader(str(path))
        blocks: list[ParsedBlock] = []
        empty_pages = 0
        for page_index, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if not text:
                empty_pages += 1
                continue
            blocks.append(ParsedBlock(text=text, block_type="pdf_page", page=page_index))
        pdf_kind = "scanned" if empty_pages == len(reader.pages) else "mixed" if empty_pages else "text"
        return _document(
            path,
            "pdf",
            self.parser_name,
            blocks,
            {"page_count": len(reader.pages), "pdf_kind": pdf_kind, "ocr_required": pdf_kind != "text"},
        )


class DocxParser(DocumentParser):
    parser_name = "docx_zip_xml"

    def parse(self, path: Path) -> ParsedDocument:
        if not zipfile.is_zipfile(path):
            return LegacyDocParser().parse(path)
        blocks: list[ParsedBlock] = []
        heading_stack: list[tuple[int, str]] = []
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml")
        root = ElementTree.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        paragraph_index = 0
        for paragraph in root.findall(".//w:p", ns):
            text = "".join(node.text or "" for node in paragraph.findall(".//w:t", ns)).strip()
            if not text:
                continue
            paragraph_index += 1
            style = _docx_paragraph_style(paragraph, ns)
            heading_level = _heading_level(style)
            if heading_level:
                heading_stack = [(level, title) for level, title in heading_stack if level < heading_level]
                heading_stack.append((heading_level, text))
            heading_path = [title for _, title in heading_stack]
            blocks.append(
                ParsedBlock(
                    text=text,
                    block_type="heading" if heading_level else "paragraph",
                    section=heading_path[-1] if heading_path else None,
                    heading_path=heading_path,
                    metadata={"paragraph_index": paragraph_index, "style": style},
                )
            )
        for table_index, table in enumerate(root.findall(".//w:tbl", ns), start=1):
            table_text = _docx_table_to_markdown(table, ns)
            if table_text:
                blocks.append(ParsedBlock(text=table_text, block_type="table", metadata={"table_index": table_index}))
        return _document(path, "docx", self.parser_name, blocks, {"paragraph_count": paragraph_index})


class LegacyDocParser(DocumentParser):
    parser_name = "doc_binary_fallback"

    def parse(self, path: Path) -> ParsedDocument:
        # .doc 是旧二进制格式；优先尝试轻量文本抽取，避免上传后直接请求失败。
        data = path.read_bytes()
        candidates = [
            data.decode("utf-16le", errors="ignore"),
            data.decode("gb18030", errors="ignore"),
            data.decode("latin-1", errors="ignore"),
        ]
        text = max((_readable_text(candidate) for candidate in candidates), key=_readability_score).strip()
        if not text:
            raise RuntimeError("无法从 .doc 文件中提取文本，请转换为 .docx 后重试。")
        return _document(
            path,
            "doc",
            self.parser_name,
            [ParsedBlock(text=text, block_type="legacy_doc_text")],
            {"legacy_format": True, "extraction": "best_effort"},
        )


class XlsxParser(DocumentParser):
    parser_name = "xlsx_structured"

    def parse(self, path: Path) -> ParsedDocument:
        try:
            return _parse_xlsx_with_openpyxl(path)
        except ImportError:
            return self._parse_with_zip_xml(path)

    def _parse_with_zip_xml(self, path: Path) -> ParsedDocument:
        blocks: list[ParsedBlock] = []
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            shared = _read_shared_strings(archive.read("xl/sharedStrings.xml")) if "xl/sharedStrings.xml" in names else []
            sheet_names = _read_sheet_names(archive.read("xl/workbook.xml")) if "xl/workbook.xml" in names else {}
            for name in sorted(names):
                if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
                    continue
                sheet_number = re.search(r"sheet(\d+)\.xml", name)
                sheet_id = sheet_number.group(1) if sheet_number else str(len(blocks) + 1)
                sheet_name = sheet_names.get(sheet_id, f"Sheet {sheet_id}")
                blocks.extend(_xlsx_sheet_blocks(archive.read(name), shared, sheet_name))
        return _document(path, "xlsx", self.parser_name, blocks, {"sheet_count": len({b.sheet_name for b in blocks})})


class CsvParser(DocumentParser):
    parser_name = "csv"

    def parse(self, path: Path) -> ParsedDocument:
        try:
            text = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            text = path.read_text(encoding="gb18030", errors="replace")
        rows = list(csv.DictReader(text.splitlines()))
        blocks = [_row_block(row, "CSV", index) for index, row in enumerate(rows, start=1)]
        if rows:
            blocks.insert(0, ParsedBlock(text=_rows_to_markdown([list(rows[0].keys()), *[list(row.values()) for row in rows]]), block_type="table", sheet_name="CSV", metadata={"columns": list(rows[0].keys())}))
        return _document(path, "csv", self.parser_name, blocks, {"row_count": len(rows)})


class PptxParser(DocumentParser):
    parser_name = "pptx_zip_xml"

    def parse(self, path: Path) -> ParsedDocument:
        blocks: list[ParsedBlock] = []
        with zipfile.ZipFile(path) as archive:
            names = sorted(name for name in archive.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name))
            notes = _pptx_notes_by_slide(archive)
            for name in names:
                match = re.search(r"slide(\d+)\.xml", name)
                slide_number = int(match.group(1)) if match else len(blocks) + 1
                texts, tables = _pptx_slide_text_and_tables(archive.read(name))
                title = texts[0] if texts else f"Slide {slide_number}"
                parts = ["\n".join(texts)] if texts else []
                parts.extend(tables)
                if notes.get(slide_number):
                    parts.append(f"Notes:\n{notes[slide_number]}")
                blocks.append(
                    ParsedBlock(
                        text="\n\n".join(parts).strip(),
                        block_type="slide",
                        slide_number=slide_number,
                        section=title,
                        heading_path=[title],
                        metadata={"slide_title": title, "table_count": len(tables)},
                    )
                )
        return _document(path, "pptx", self.parser_name, blocks, {"slide_count": len(blocks)})


class HtmlDocumentParser(DocumentParser):
    parser_name = "html_stdlib"

    def parse(self, path: Path) -> ParsedDocument:
        raw = path.read_text(encoding="utf-8", errors="replace")
        parser = _ReadableHtmlParser()
        parser.feed(raw)
        blocks = [
            ParsedBlock(
                text=section_text,
                block_type="html_section",
                section=heading_path[-1] if heading_path else parser.title or path.stem,
                heading_path=heading_path,
                metadata={"title": parser.title or path.stem, "url": str(path)},
            )
            for section_text, heading_path in parser.sections()
            if section_text.strip()
        ]
        if not blocks and parser.text:
            blocks = [ParsedBlock(text=parser.text, block_type="html_body", metadata={"title": parser.title or path.stem, "url": str(path)})]
        return _document(path, "html", self.parser_name, blocks, {"title": parser.title or path.stem, "url": str(path)})


class ImageOcrParser(DocumentParser):
    parser_name = "image_ocr"

    def parse(self, path: Path) -> ParsedDocument:
        metadata: dict[str, object] = {"image_name": path.name, "ocr_confidence": None}
        try:
            from PIL import Image
            import pytesseract
        except ImportError:
            metadata["ocr_available"] = False
            text = "OCR dependencies are not installed. Install Pillow and pytesseract, and configure Tesseract OCR."
            return _document(path, "image", self.parser_name, [ParsedBlock(text=text, block_type="ocr_unavailable", metadata=metadata)], metadata)

        try:
            image = Image.open(path)
            data = pytesseract.image_to_data(image, lang="chi_sim+eng", output_type=pytesseract.Output.DICT)
            words = [word.strip() for word in data.get("text", []) if word and word.strip()]
            confidences = [float(value) for value in data.get("conf", []) if _is_confidence(value)]
            metadata["ocr_available"] = True
            metadata["ocr_confidence"] = round(sum(confidences) / len(confidences), 4) if confidences else None
            text = " ".join(words).strip()
        except Exception as exc:
            metadata["ocr_available"] = False
            metadata["ocr_error"] = str(exc)
            text = f"OCR failed for {path.name}: {exc}"
        return _document(path, "image", self.parser_name, [ParsedBlock(text=text, block_type="ocr_text", metadata=metadata)], metadata)


def _document(
    path: Path,
    file_type: str,
    parser: str,
    blocks: list[ParsedBlock],
    metadata: dict[str, object] | None = None,
) -> ParsedDocument:
    return ParsedDocument(
        source=str(path),
        file_name=path.name,
        file_type=file_type,
        title=path.stem,
        parser=parser,
        blocks=blocks,
        metadata=metadata or {},
    )


def _readable_text(text: str) -> str:
    lines = []
    for line in text.splitlines():
        cleaned = "".join(char if char.isprintable() else " " for char in line)
        cleaned = " ".join(cleaned.split())
        if len(cleaned) >= 2:
            lines.append(cleaned)
    return "\n".join(lines)


def _readability_score(text: str) -> int:
    cjk_count = sum(1 for char in text if "\u4e00" <= char <= "\u9fff")
    ascii_word_count = sum(1 for char in text if char.isascii() and char.isalnum())
    return (cjk_count * 8) + ascii_word_count + len(text)


def _try_complex_pdf_backend(path: Path, backend: str) -> ParsedDocument | None:
    candidates = ["docling", "unstructured", "mineru"] if backend == "auto" else [backend]
    errors: dict[str, str] = {}
    for candidate in candidates:
        try:
            if candidate == "docling":
                return _parse_pdf_with_docling(path)
            if candidate == "unstructured":
                return _parse_pdf_with_unstructured(path)
            if candidate == "mineru":
                return _parse_pdf_with_mineru(path)
        except Exception as exc:
            errors[candidate] = str(exc)
    if backend != "auto" and errors:
        raise RuntimeError(f"PDF parser backend {backend!r} failed: {errors[backend]}") from None
    return None


def _parse_xlsx_with_openpyxl(path: Path) -> ParsedDocument:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    blocks: list[ParsedBlock] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        headers = _normalize_headers(rows[0]) if rows else []
        if headers:
            table_preview = _rows_to_markdown([headers, *[_row_values(row, len(headers)) for row in rows[1:8]]])
            if table_preview:
                blocks.append(
                    ParsedBlock(
                        text=table_preview,
                        block_type="table",
                        sheet_name=sheet.title,
                        metadata={"columns": headers, "sheet_name": sheet.title},
                    )
                )
        for row_index, row in enumerate(rows[1:] if headers else rows, start=2 if headers else 1):
            values = _row_values(row, len(headers) if headers else len(row))
            if not any(values):
                continue
            row_map = {
                headers[index] if index < len(headers) and headers[index] else f"字段{index + 1}": values[index]
                for index in range(len(values))
                if values[index]
            }
            text = _structured_row_text(row_map)
            if not text:
                continue
            blocks.append(
                ParsedBlock(
                    text=text,
                    block_type="table_row",
                    sheet_name=sheet.title,
                    row_index=row_index,
                    section=sheet.title,
                    heading_path=[sheet.title],
                    metadata={
                        "columns": list(row_map.keys()),
                        "row_json": row_map,
                        "business_scene": _first_value(row_map, ("业务场景", "适用场景", "场景", "scene")),
                        "topic": _first_value(row_map, ("主题", "问题", "话术类型", "topic")),
                        "risk_level": _infer_risk_level_from_row(row_map),
                        "applicable_stage": _first_value(row_map, ("适用阶段", "账龄", "阶段", "stage")),
                    },
                )
            )
    return _document(path, "xlsx", "openpyxl_structured", blocks, {"sheet_count": len(workbook.worksheets)})


def _parse_pdf_with_docling(path: Path) -> ParsedDocument:
    module = importlib.import_module("docling.document_converter")
    converter_cls = getattr(module, "DocumentConverter")
    result = converter_cls().convert(str(path))
    document = getattr(result, "document", result)
    markdown = document.export_to_markdown() if hasattr(document, "export_to_markdown") else str(document)
    blocks = [ParsedBlock(text=markdown, block_type="pdf_layout", metadata={"layout_backend": "docling"})]
    return _document(path, "pdf", "docling", blocks, {"pdf_kind": "complex", "layout_backend": "docling"})


def _parse_pdf_with_unstructured(path: Path) -> ParsedDocument:
    module = importlib.import_module("unstructured.partition.pdf")
    partition_pdf = getattr(module, "partition_pdf")
    elements = partition_pdf(filename=str(path), infer_table_structure=True, strategy="hi_res")
    blocks = []
    for index, element in enumerate(elements):
        text = str(element).strip()
        if not text:
            continue
        metadata = getattr(element, "metadata", None)
        page = getattr(metadata, "page_number", None) if metadata is not None else None
        category = getattr(element, "category", element.__class__.__name__)
        blocks.append(
            ParsedBlock(
                text=text,
                block_type=f"pdf_{str(category).lower()}",
                page=page,
                metadata={"layout_backend": "unstructured", "element_index": index, "category": str(category)},
            )
        )
    return _document(path, "pdf", "unstructured", blocks, {"pdf_kind": "complex", "layout_backend": "unstructured"})


def _parse_pdf_with_mineru(path: Path) -> ParsedDocument:
    module = importlib.import_module("mineru")
    parse_pdf = getattr(module, "parse_pdf")
    result = parse_pdf(str(path))
    markdown = result.get("markdown", "") if isinstance(result, dict) else str(result)
    blocks = [ParsedBlock(text=markdown, block_type="pdf_layout", metadata={"layout_backend": "mineru"})]
    return _document(path, "pdf", "mineru", blocks, {"pdf_kind": "complex", "layout_backend": "mineru"})


def _docx_paragraph_style(paragraph: ElementTree.Element, ns: dict[str, str]) -> str:
    style = paragraph.find("./w:pPr/w:pStyle", ns)
    return style.attrib.get(f"{{{ns['w']}}}val", "") if style is not None else ""


def _heading_level(style: str) -> int | None:
    match = re.search(r"Heading(\d+)", style, re.IGNORECASE)
    return int(match.group(1)) if match else None


def _docx_table_to_markdown(table: ElementTree.Element, ns: dict[str, str]) -> str:
    rows: list[list[str]] = []
    for row in table.findall(".//w:tr", ns):
        values = ["".join(node.text or "" for node in cell.findall(".//w:t", ns)).strip() for cell in row.findall("./w:tc", ns)]
        if any(values):
            rows.append(values)
    return _rows_to_markdown(rows)


def _read_shared_strings(xml: bytes) -> list[str]:
    root = ElementTree.fromstring(xml)
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return ["".join(node.text or "" for node in item.findall(".//x:t", ns)).strip() for item in root.findall(".//x:si", ns)]


def _read_sheet_names(xml: bytes) -> dict[str, str]:
    root = ElementTree.fromstring(xml)
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return {str(index): html.unescape(sheet.attrib["name"]) for index, sheet in enumerate(root.findall(".//x:sheet", ns), start=1) if sheet.attrib.get("name")}


def _xlsx_sheet_blocks(xml: bytes, shared: list[str], sheet_name: str) -> list[ParsedBlock]:
    root = ElementTree.fromstring(xml)
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rows: list[list[str]] = []
    for row in root.findall(".//x:sheetData/x:row", ns):
        values = [_xlsx_cell(cell, shared, ns) for cell in row.findall("x:c", ns)]
        if any(values):
            rows.append(values)
    if not rows:
        return []
    headers = rows[0]
    blocks = [ParsedBlock(text=_rows_to_markdown(rows), block_type="table", sheet_name=sheet_name, metadata={"columns": headers})]
    for index, values in enumerate(rows[1:], start=2):
        row = {headers[i] if i < len(headers) and headers[i] else f"column_{i + 1}": value for i, value in enumerate(values)}
        blocks.append(_row_block(row, sheet_name, index))
    return blocks


def _normalize_headers(row: tuple[object, ...]) -> list[str]:
    return [clean_text(str(value)) if value is not None else "" for value in row]


def _row_values(row: tuple[object, ...], width: int) -> list[str]:
    values = [clean_text(str(value)) if value is not None else "" for value in row]
    if len(values) < width:
        values.extend([""] * (width - len(values)))
    return values[:width]


def _structured_row_text(row: dict[str, str]) -> str:
    priority_keys = ("问题", "标准答案", "答案", "话术", "适用场景", "业务场景", "合规要求", "禁止事项", "注意事项")
    ordered_keys = [key for key in priority_keys if key in row]
    ordered_keys.extend(key for key in row if key not in ordered_keys)
    lines = [f"{key}: {row[key]}" for key in ordered_keys if row.get(key)]
    return clean_text("\n".join(lines))


def _first_value(row: dict[str, str], names: tuple[str, ...]) -> str:
    lowered = {key.lower(): value for key, value in row.items()}
    for name in names:
        if name in row and row[name]:
            return row[name]
        if name.lower() in lowered and lowered[name.lower()]:
            return lowered[name.lower()]
    return ""


def _infer_risk_level_from_row(row: dict[str, str]) -> str:
    text = " ".join(row.values())
    high_markers = ("法务", "诉讼", "起诉", "律师函", "威胁", "报警", "征信", "承诺撤案", "敏感词")
    medium_markers = ("投诉", "不满", "拒绝", "联系人", "合规", "承诺")
    if any(marker in text for marker in high_markers):
        return "high"
    if any(marker in text for marker in medium_markers):
        return "medium"
    return "low"


def _xlsx_cell(cell: ElementTree.Element, shared: list[str], ns: dict[str, str]) -> str:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//x:t", ns)).strip()
    value_node = cell.find("x:v", ns)
    if value_node is None or value_node.text is None:
        return ""
    value = value_node.text.strip()
    if cell_type == "s":
        try:
            return shared[int(value)]
        except (IndexError, ValueError):
            return value
    return value


def _row_block(row: dict[str, str], sheet_name: str, row_index: int) -> ParsedBlock:
    natural = "，".join(f"{key}为{value}" for key, value in row.items() if value)
    return ParsedBlock(
        text=natural,
        block_type="table_row",
        sheet_name=sheet_name,
        row_index=row_index,
        metadata={"columns": list(row.keys()), "row_json": row},
    )


def _rows_to_markdown(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    header = normalized[0]
    separator = ["---"] * width
    body = normalized[1:]
    return "\n".join(["| " + " | ".join(header) + " |", "| " + " | ".join(separator) + " |", *["| " + " | ".join(row) + " |" for row in body]])


def _pptx_slide_text_and_tables(xml: bytes) -> tuple[list[str], list[str]]:
    root = ElementTree.fromstring(xml)
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    texts = []
    for paragraph in root.findall(".//a:p", ns):
        line = "".join(node.text or "" for node in paragraph.findall(".//a:t", ns)).strip()
        if line:
            texts.append(line)
    tables = []
    for table in root.findall(".//a:tbl", ns):
        rows = []
        for row in table.findall(".//a:tr", ns):
            values = []
            for cell in row.findall(".//a:tc", ns):
                values.append("".join(node.text or "" for node in cell.findall(".//a:t", ns)).strip())
            if any(values):
                rows.append(values)
        table_text = _rows_to_markdown(rows)
        if table_text:
            tables.append(table_text)
    return texts, tables


def _pptx_notes_by_slide(archive: zipfile.ZipFile) -> dict[int, str]:
    notes: dict[int, str] = {}
    for name in archive.namelist():
        match = re.fullmatch(r"ppt/notesSlides/notesSlide(\d+)\.xml", name)
        if not match:
            continue
        root = ElementTree.fromstring(archive.read(name))
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        lines = []
        for paragraph in root.findall(".//a:p", ns):
            line = "".join(node.text or "" for node in paragraph.findall(".//a:t", ns)).strip()
            if line:
                lines.append(line)
        if lines:
            notes[int(match.group(1))] = "\n".join(lines)
    return notes


class _ReadableHtmlParser(HTMLParser):
    skip_tags = {"script", "style", "nav", "footer", "aside", "noscript"}

    def __init__(self) -> None:
        super().__init__()
        self.title = ""
        self.text_parts: list[str] = []
        self.heading_stack: list[tuple[int, str]] = []
        self.records: list[tuple[str, list[str]]] = []
        self.current_tag = ""
        self.skip_depth = 0

    @property
    def text(self) -> str:
        return "\n".join(part for part in self.text_parts if part.strip()).strip()

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        self.current_tag = tag.lower()
        if self.current_tag in self.skip_tags:
            self.skip_depth += 1

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in self.skip_tags and self.skip_depth:
            self.skip_depth -= 1
        self.current_tag = ""

    def handle_data(self, data: str) -> None:
        if self.skip_depth:
            return
        text = " ".join(html.unescape(data).split())
        if not text:
            return
        if self.current_tag == "title":
            self.title = text
            return
        if self.current_tag in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            level = int(self.current_tag[1])
            self.heading_stack = [(item_level, item_title) for item_level, item_title in self.heading_stack if item_level < level]
            self.heading_stack.append((level, text))
            self.records.append((f"{'#' * min(level, 6)} {text}", [item_title for _, item_title in self.heading_stack]))
            return
        if self.current_tag in {"p", "li", "td", "th", "div", "article", "section", "main"}:
            heading_path = [item_title for _, item_title in self.heading_stack]
            self.records.append((text, heading_path))
            self.text_parts.append(text)

    def sections(self) -> list[tuple[str, list[str]]]:
        if not self.records:
            return []
        sections: list[tuple[str, list[str]]] = []
        current_lines: list[str] = []
        current_heading: list[str] = []
        for text, heading_path in self.records:
            if text.startswith("#") and current_lines:
                sections.append(("\n".join(current_lines), current_heading))
                current_lines = []
            current_heading = heading_path or current_heading
            current_lines.append(text)
        if current_lines:
            sections.append(("\n".join(current_lines), current_heading))
        return sections


def _is_confidence(value: object) -> bool:
    try:
        return float(value) >= 0
    except (TypeError, ValueError):
        return False
